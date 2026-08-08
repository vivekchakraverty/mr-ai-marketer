"""Tracker Studio — storage for the Manage screen's two workbooks.

The screen is a faithful rebuild of two spreadsheets the user works in:
"Social Media Ads Tracker" (8 sheets) and "Marketing Influencer Outreach
Tracker" (3 sheets). This router only stores and serves their *input* cells.
Every derived column — the SUMIFS/SUMIF/IFERROR/FILTER columns, the dashboard
KPIs, the status roll-ups — is recomputed in the renderer on each keystroke, the
way a spreadsheet recalculates. Persisting computed values would let them drift
from their inputs, which is exactly the failure mode a spreadsheet doesn't have.

Dates are stored as Excel serial numbers (days since 1899-12-30), not ISO
strings, because that is what the formulas operate on: `MAX(H4-TODAY(),0)` and
the `>=start` / `<next month` window comparisons are integer arithmetic in the
workbook, and keeping the same representation keeps them exact.
"""
from __future__ import annotations

from typing import Any

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

from .. import db

router = APIRouter(prefix="/tracker", tags=["tracker"])


def _rows(letters: str, records: list[list[Any]]) -> list[dict]:
    """Zip positional seed rows onto their spreadsheet column letters.

    The source sheets are addressed by letter and skip columns (the influencer
    tracker leaves F, J, N and R blank as spacers), so rows travel as
    {"B": ..., "C": ...} rather than as positional arrays.
    """
    keys = letters.split()
    return [dict(zip(keys, record)) for record in records]


# --- Seed data -------------------------------------------------------------
# Transcribed from the two source workbooks, values verbatim. A fresh install
# opens on the same data the spreadsheets ship with, and "Reset data" restores
# exactly this.

_DAILY = _rows(
    "A B C D E F G H I J K L M N W",
    [
        [46204, "Credible Sheets", "Facebook", "CMP-001", "Payroll Web App Leads", "Business Owners PH", "Static - Payroll Pain Points", 1850, 44200, 31800, 1240, 86, 12, 37200, "Strong opening day"],
        [46205, "Credible Sheets", "Facebook", "CMP-001", "Payroll Web App Leads", "Business Owners PH", "Video - Save Payroll Time", 2120, 51800, 36600, 1518, 102, 16, 49600, "Video outperformed static"],
        [46206, "Credible Sheets", "Instagram", "CMP-002", "POS Tracker Awareness", "Retail Owners", "Carousel - POS Features", 1400, 38600, 29400, 910, 44, 7, 19600, "Good CTR"],
        [46207, "Client A", "TikTok", "CMP-003", "Skin Care Sales", "Beauty Interest 18-34", "UGC - Before After", 2750, 112000, 84900, 2380, 121, 23, 80500, "High reach"],
        [46208, "Client A", "TikTok", "CMP-003", "Skin Care Sales", "Beauty Interest 18-34", "UGC - Testimonial", 2480, 98500, 76200, 2010, 113, 19, 66500, "Stable CPA"],
        [46209, "Client B", "Google Ads", "CMP-004", "Accounting Search Leads", "High Intent Keywords", "Responsive Search Ad A", 3200, 22400, 19700, 1780, 94, 18, 99000, "Best ROAS"],
        [46210, "Client B", "Google Ads", "CMP-004", "Accounting Search Leads", "Competitor Keywords", "Responsive Search Ad B", 2980, 20700, 18500, 1510, 71, 11, 60500, "Higher CPC"],
        [46211, "Client C", "Facebook", "CMP-005", "Property Inquiry Leads", "OFW Audience", "Video - Condo Tour", 4300, 74600, 52800, 1920, 138, 9, 270000, "High-value conversions"],
        [46212, "Client C", "Instagram", "CMP-005", "Property Inquiry Leads", "Local Professionals", "Reel - Amenities", 3650, 68800, 49700, 1760, 119, 6, 180000, "Retarget viewers"],
        [46213, "Client D", "LinkedIn", "CMP-006", "B2B Demo Campaign", "Managers PH", "Document Ad - Case Study", 5100, 31200, 24600, 840, 52, 8, 208000, "Qualified leads"],
        [46214, "Credible Sheets", "YouTube", "CMP-007", "Web App Demo Views", "SME Owners", "15-sec Demo Video", 1700, 58400, 42100, 1030, 38, 5, 14200, "Low-cost reach"],
        [46215, "Client A", "Facebook", "CMP-008", "Retargeting Sales", "Website Visitors 30D", "Dynamic Product Ad", 2250, 40800, 20900, 1380, 79, 21, 73500, "Excellent frequency"],
    ],
)

_CAMPAIGNS = _rows(
    "A B C D E F G H I J K L M N O P",
    [
        ["CMP-001", "Credible Sheets", "Facebook", "Credible Sheets Ads", "Payroll Web App Leads", "Leads", 46204, 46234, "Monthly", 45000, 1500, "Active", "Edzel", "SME and business owners", "https://example.com/payroll", "utm_campaign=payroll"],
        ["CMP-002", "Credible Sheets", "Instagram", "Credible Sheets Ads", "POS Tracker Awareness", "Awareness", 46204, 46234, "Monthly", 25000, 850, "Active", "Marketing Team", "Retail and restaurant owners", "https://example.com/pos", "utm_campaign=pos"],
        ["CMP-003", "Client A", "TikTok", "Client A Main", "Skin Care Sales", "Sales", 46204, 46223, "Lifetime", 65000, 3250, "Active", "Media Buyer", "Beauty interest 18-34", "https://example.com/skincare", "utm_campaign=skin"],
        ["CMP-004", "Client B", "Google Ads", "Client B Search", "Accounting Search Leads", "Leads", 46204, 46234, "Monthly", 90000, 3000, "Active", "Marketing Team", "High-intent business searches", "https://example.com/accounting", "utm_campaign=accounting"],
        ["CMP-005", "Client C", "Facebook", "Client C Property", "Property Inquiry Leads", "Leads", 46208, 46249, "Lifetime", 160000, 4000, "Active", "Media Buyer", "OFWs and professionals", "https://example.com/property", "utm_campaign=property"],
        ["CMP-006", "Client D", "LinkedIn", "Client D B2B", "B2B Demo Campaign", "Leads", 46211, 46242, "Lifetime", 120000, 3750, "Active", "Edzel", "Managers and decision makers", "https://example.com/demo", "utm_campaign=b2b"],
        ["CMP-007", "Credible Sheets", "YouTube", "Credible Sheets Video", "Web App Demo Views", "Traffic", 46213, 46234, "Lifetime", 30000, 1400, "Active", "Designer", "SME owners", "https://example.com/demo-video", "utm_campaign=youtube"],
        ["CMP-008", "Client A", "Facebook", "Client A Main", "Retargeting Sales", "Sales", 46213, 46234, "Lifetime", 35000, 1600, "Active", "Media Buyer", "Website visitors 30 days", "https://example.com/shop", "utm_campaign=retarget"],
    ],
)

# Columns I (Asset URL) and V (Notes) are blank in the source workbook.
_CREATIVES = _rows(
    "A B C D E F G H I J K V",
    [
        ["CR-001", "Credible Sheets", "Facebook", "CMP-001", "Static - Payroll Pain Points", "Image", "Stop wasting hours on payroll", "Learn More", "", 46204, "Active", ""],
        ["CR-002", "Credible Sheets", "Facebook", "CMP-001", "Video - Save Payroll Time", "Video", "Finish payroll in minutes", "Get Quote", "", 46205, "Active", ""],
        ["CR-003", "Credible Sheets", "Instagram", "CMP-002", "Carousel - POS Features", "Carousel", "Everything your store needs", "Learn More", "", 46206, "Active", ""],
        ["CR-004", "Client A", "TikTok", "CMP-003", "UGC - Before After", "Video", "Real results in 14 days", "Shop Now", "", 46207, "Active", ""],
        ["CR-005", "Client A", "TikTok", "CMP-003", "UGC - Testimonial", "Video", "Why customers keep reordering", "Shop Now", "", 46208, "Active", ""],
        ["CR-006", "Client B", "Google Ads", "CMP-004", "Responsive Search Ad A", "Search Ad", "Affordable accounting support", "Get Quote", "", 46209, "Active", ""],
        ["CR-007", "Client C", "Facebook", "CMP-005", "Video - Condo Tour", "Video", "Own a home while abroad", "Book Now", "", 46211, "Active", ""],
        ["CR-008", "Client D", "LinkedIn", "CMP-006", "Document Ad - Case Study", "Carousel", "See how we improved operations", "Download", "", 46213, "Active", ""],
    ],
)

_LEADS = _rows(
    "A B C D E F G H I J K L M N",
    [
        ["LD-001", 46204, "Credible Sheets", "Facebook", "CMP-001", "Juan Dela Cruz", "0917-000-1001", "Qualified", "Sales Team", 46206, 46208, 3100, "Converted", "Purchased payroll system"],
        ["LD-002", 46205, "Credible Sheets", "Facebook", "CMP-001", "Maria Santos", "maria@example.com", "Contacted", "Sales Team", 46207, "", 0, "Open", "Needs demo"],
        ["LD-003", 46207, "Client A", "TikTok", "CMP-003", "Angela Reyes", "0917-000-1003", "Won", "Sales Team", 46208, 46209, 3500, "Converted", "Online order"],
        ["LD-004", 46209, "Client B", "Google Ads", "CMP-004", "ABC Trading", "abc@example.com", "Proposal", "Edzel", 46211, "", 0, "Qualified", "Sent quotation"],
        ["LD-005", 46211, "Client C", "Facebook", "CMP-005", "Roberto Garcia", "0917-000-1005", "Qualified", "Sales Team", 46213, "", 0, "Open", "OFW property inquiry"],
        ["LD-006", 46213, "Client D", "LinkedIn", "CMP-006", "Northstar Inc.", "info@northstar.example", "Won", "Edzel", 46214, 46217, 26000, "Converted", "Demo package"],
    ],
)

_BUDGET = _rows(
    "A B C D E F",
    [
        [46204, "Credible Sheets", 70000, 300, 45, 140000],
        [46204, "Client A", 100000, 350, 65, 220000],
        [46204, "Client B", 90000, 220, 40, 250000],
        [46204, "Client C", 160000, 250, 20, 600000],
        [46204, "Client D", 120000, 100, 15, 350000],
    ],
)

_INFLUENCERS = _rows(
    "B C D E F G H I J K L M N O",
    [
        ["John Influencer", "john@email.com", "Instagram", 50000, "Sponsored Post", 44936, "Pending", 44941, "In Progress", "Sent collaboration proposal including product details, brand guidelines, and content expectations.", 44956, 44962, "Payment sent via PayPal on 02/08/2023.", "Track engagement metrics, reach, and impressions on sponsored post."],
        ["Jane Blogger", "jane@email.com", "YouTube", 100000, "Product Review", 44938, "Declined", "-", "Not Collaborating", "Declined due to current content schedule; suggested a possible collaboration in the future.", 44957, 44963, "-", "-"],
        ["Alex Vlogger", "alex@email.com", "Instagram", 80000, "Giveaway", 44941, "Accepted", 44946, "Collaborating", "Accepted collaboration terms, provided shipping address for giveaway items.", 44958, 44964, "-", "Track giveaway entries and engagement."],
    ],
)

# The Settings sheet's dropdown source columns (A..J), in sheet order. These feed
# every list-validated column in the Ads workbook.
_LISTS = [
    {"key": "platforms", "name": "Platforms", "items": ["All", "Facebook", "Instagram", "TikTok", "Google Ads", "YouTube", "LinkedIn"]},
    {"key": "objectives", "name": "Objectives", "items": ["Awareness", "Traffic", "Engagement", "Leads", "Sales", "App Promotion", "Messages", "Value"]},
    {"key": "campaignStatus", "name": "Campaign Status", "items": ["Planning", "Active", "Paused", "Completed", "Cancelled"]},
    {"key": "budgetType", "name": "Budget Type", "items": ["Daily", "Lifetime", "Monthly", "Project"]},
    {"key": "owners", "name": "Owners", "items": ["Edzel", "Marketing Team", "Media Buyer", "Designer", "Sales Team"]},
    {"key": "clients", "name": "Clients", "items": ["Credible Sheets", "Client A", "Client B", "Client C", "Client D"]},
    {"key": "creativeFormat", "name": "Creative Format", "items": ["Image", "Video", "Carousel", "Reel", "Story", "Search Ad", "Display Ad"]},
    {"key": "cta", "name": "CTA", "items": ["Learn More", "Shop Now", "Sign Up", "Book Now", "Get Quote", "Contact Us", "Download"]},
    {"key": "leadStage", "name": "Lead Stage", "items": ["New", "Contacted", "Qualified", "Proposal", "Won", "Lost"]},
    {"key": "leadStatus", "name": "Lead Status", "items": ["Open", "Qualified", "Converted", "Lost", "Closed"]},
]

# Settings!A12:B16 — the "Tracker Defaults" block.
_DEFAULTS = {
    "defaultStartDate": 46204,
    "defaultEndDate": 46234,
    "currency": "PHP",
    "timeZone": "Asia/Manila",
    "notes": "Add or replace list values above before using the tracker.",
}

# Dashboard!B4/D4/F4/H4 — the report filter cells, and the influencer tracker's
# own platform cell (F4). Stored so a reopened screen lands where it was left.
_FILTERS = {
    "start": 46204,
    "end": 46234,
    "platform": "All",
    "client": "Client A",
    "influencerPlatform": "Instagram",
}

# The influencer workbook's Setup!B5:B7 list, and its footer line (B9).
_SETUP = ["Instagram", "YouTube", "TikTok"]
_COMPANY_LINE = "Company Address   I   Company Email   I   Company Website   I   Company Number"


def _seed() -> dict:
    return {
        "daily": _DAILY,
        "campaigns": _CAMPAIGNS,
        "creatives": _CREATIVES,
        "leads": _LEADS,
        "budget": _BUDGET,
        "influencers": _INFLUENCERS,
        "lists": _LISTS,
        "defaults": _DEFAULTS,
        "filters": _FILTERS,
        "setup": _SETUP,
        "companyLine": _COMPANY_LINE,
    }


class Workbooks(BaseModel):
    daily: list[dict]
    campaigns: list[dict]
    creatives: list[dict]
    leads: list[dict]
    budget: list[dict]
    influencers: list[dict]
    lists: list[dict]
    defaults: dict
    filters: dict
    setup: list[str]
    companyLine: str


def _current() -> dict:
    """Stored state, with any missing sheet filled in from the seed.

    Per-key rather than all-or-nothing so that a workbook gaining a sheet in a
    later version doesn't blank out the sheets the user has already edited.
    """
    stored = db.get_tracker_docs()
    seeded = _seed()
    seeded.update({key: value for key, value in stored.items() if key in seeded})
    return seeded


@router.get("/workbooks", response_model=Workbooks)
def get_workbooks() -> Workbooks:
    return Workbooks(**_current())


@router.put("/workbooks", response_model=Workbooks)
def put_workbooks(body: Workbooks) -> Workbooks:
    db.put_tracker_docs(body.model_dump())
    return body


@router.post("/reset", response_model=Workbooks)
def reset_workbooks() -> Workbooks:
    """Restore both workbooks to the data the source spreadsheets ship with."""
    db.clear_tracker_docs()
    return Workbooks(**_seed())


# --- Export ----------------------------------------------------------------
#
# The export takes its rows from the *renderer*, not from storage, and that is deliberate.
# This router only ever holds input cells — every SUMIFS total, every dashboard KPI, every
# status roll-up is recomputed in the browser (see the module docstring). An export built
# from stored state would therefore be missing exactly the columns people open a tracker to
# read. So the screen sends what it is displaying, headers and computed values included, and
# this turns it into a file.


class ExportSheet(BaseModel):
    name: str
    columns: list[str]
    rows: list[list[str]]


class ExportRequest(BaseModel):
    format: str = "xlsx"  # "xlsx" | "csv"
    workbook: str = "tracker"
    sheets: list[ExportSheet]


def _safe(name: str, fallback: str) -> str:
    cleaned = "".join(c for c in name if c.isalnum() or c in " -_").strip()
    return cleaned or fallback


@router.post("/export")
def export_workbook(body: ExportRequest) -> dict:
    """Write the supplied sheets to disk and return where they landed.

    Files rather than a streamed download: this app hands finished artefacts to the OS the
    same way DocuMaker and the Blog Writer do, so the user gets a real file in a real folder
    that opens in Excel, not a browser download that vanishes into a temp directory.
    """
    from datetime import datetime, timezone
    from pathlib import Path

    from .. import config

    sheets = [s for s in body.sheets if s.columns]
    if not sheets:
        raise HTTPException(status_code=400, detail="Nothing to export.")

    out_dir: Path = config.OUTPUTS_DIR / "tracker"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    base = f"{_safe(body.workbook, 'tracker')}-{stamp}"

    if body.format == "csv":
        # One file per sheet. A single CSV cannot hold several tables without inventing a
        # convention that no spreadsheet program reads back.
        import csv

        folder = out_dir / base
        folder.mkdir(parents=True, exist_ok=True)
        written = []
        for sheet in sheets:
            path = folder / f"{_safe(sheet.name, 'sheet')}.csv"
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                # utf-8-sig: Excel on Windows reads a plain UTF-8 CSV as the local codepage
                # and mangles every non-ASCII character. The BOM is what makes it read UTF-8.
                writer = csv.writer(handle)
                writer.writerow(sheet.columns)
                writer.writerows(sheet.rows)
            written.append(path.name)
        return {"path": str(folder), "files": written, "format": "csv"}

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    book = Workbook()
    book.remove(book.active)
    for sheet in sheets:
        # Excel sheet names are capped at 31 characters and reject []:*?/\
        title = _safe(sheet.name, "Sheet")[:31]
        ws = book.create_sheet(title)
        ws.append(sheet.columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in sheet.rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        for i, header in enumerate(sheet.columns, start=1):
            longest = max([len(str(header))] + [len(str(r[i - 1])) for r in sheet.rows if i <= len(r)] or [0])
            ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 48)

    path = out_dir / f"{base}.xlsx"
    book.save(path)
    return {"path": str(path), "files": [path.name], "format": "xlsx"}
